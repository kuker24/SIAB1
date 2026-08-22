#!/usr/bin/env python3
"""Destructive year-rollover cleanup for SIAB1.

Default mode is dry-run. Apply mode permanently deletes finished + draft exams,
removes class-12 students, and syncs class 10/11 accounts from the official
Excel roster.

This script never prints full DATABASE_URL credentials.
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import random
import re
import secrets
import shutil
import stat
import string
import sys
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Sequence
from urllib.parse import urlsplit

PRODUCTION_URL_PATTERNS = ("103.175.218.56", "man1rokanhulu.cloud", "adminujian")
DEFAULT_EXCEL = Path(
    "/home/fahmiagent/Downloads/LAB GITHUB/LAB FINAL/SIAB1/SIAB1 akun/"
    "PEMBAGIAN KELAS TP 2026-2027 terbaru Agustus.xlsx"
)
DEFAULT_REPORT_DIR = Path("MT/reports")
DEFAULT_MT_ARCHIVES = (
    Path("MT/MT2026-08-10"),
    Path("MT/MT2026-08-10-teacher-only-backup"),
)
ROSTER_SHEET_ALIASES = {
    "XA": "X A",
    "XB": "X B",
    "XC": "X C",
    "XD": "X D",
    "XE": "X E",
    "XF": "X F",
    "XIA": "XI A",
    "XIB": "XI B",
    "XIC": "XI C",
    "XID": "XI D",
    "XIE": "XI E",
}
# XII sheets are intentionally ignored for re-import.
IGNORED_SHEET_PREFIXES = ("XII",)

EXAM_TARGET_SQL = """
SELECT
    e.id AS exam_id,
    e.title,
    e.is_published,
    e.is_deleted,
    e.start_time,
    e.end_time,
    CASE
        WHEN e.is_published IS NOT TRUE THEN 'draft'
        WHEN e.end_time IS NOT NULL AND e.end_time <= :captured_at THEN 'finished'
        WHEN e.is_published IS TRUE
             AND e.start_time IS NOT NULL
             AND e.end_time IS NOT NULL
             AND e.start_time <= :captured_at
             AND e.end_time > :captured_at THEN 'active'
        ELSE 'other_published'
    END AS lifecycle
FROM exams e
ORDER BY e.id
"""

EXAM_CHILD_COUNTS_SQL = """
SELECT
    (SELECT COUNT(*) FROM exam_sessions es WHERE es.exam_id = ANY(:exam_ids)) AS session_count,
    (SELECT COUNT(*) FROM answers a
        WHERE a.session_id IN (SELECT id FROM exam_sessions WHERE exam_id = ANY(:exam_ids))
    ) AS answer_count,
    (SELECT COUNT(*) FROM exam_logs el
        WHERE el.session_id IN (SELECT id FROM exam_sessions WHERE exam_id = ANY(:exam_ids))
    ) AS log_count,
    (SELECT COUNT(*) FROM questions q WHERE q.exam_id = ANY(:exam_ids)) AS question_count
"""

STUDENT_SQL = """
SELECT
    id,
    username,
    full_name,
    role,
    student_class,
    is_active
FROM users
WHERE lower(trim(role)) = 'student'
ORDER BY id
"""


class CleanupError(RuntimeError):
    """Raised when cleanup cannot run safely."""


@dataclass
class RosterStudent:
    full_name: str
    student_class: str
    source_sheet: str
    row_number: int
    nisn: Optional[str] = None

    @property
    def normalized_name(self) -> str:
        return normalize_person_name(self.full_name)


@dataclass
class DbStudent:
    id: int
    username: str
    full_name: str
    student_class: Optional[str]
    is_active: bool

    @property
    def normalized_name(self) -> str:
        return normalize_person_name(self.full_name)

    @property
    def canonical_class(self) -> Optional[str]:
        return normalize_class_name(self.student_class)

    @property
    def grade(self) -> Optional[str]:
        return class_grade(self.student_class)


@dataclass
class MatchPlan:
    matched_updates: list[dict[str, Any]] = field(default_factory=list)
    create_new: list[dict[str, Any]] = field(default_factory=list)
    ambiguous: list[dict[str, Any]] = field(default_factory=list)
    deactivate: list[dict[str, Any]] = field(default_factory=list)
    unchanged: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class ExamPlan:
    draft_ids: list[int] = field(default_factory=list)
    finished_ids: list[int] = field(default_factory=list)
    active_ids: list[int] = field(default_factory=list)
    other_ids: list[int] = field(default_factory=list)
    target_ids: list[int] = field(default_factory=list)
    child_counts: dict[str, int] = field(default_factory=dict)


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--excel",
        type=Path,
        default=DEFAULT_EXCEL,
        help="Path ke file pembagian kelas X/XI.",
    )
    parser.add_argument(
        "--report-dir",
        type=Path,
        default=DEFAULT_REPORT_DIR,
        help="Folder laporan privat (default MT/reports).",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        default=True,
        help="Hitung dan tulis laporan saja (default).",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Jalankan perubahan destruktif ke database.",
    )
    parser.add_argument(
        "--allow-production-write",
        action="store_true",
        help="Izinkan write ke DATABASE_URL produksi.",
    )
    parser.add_argument(
        "--i-understand-destructive",
        action="store_true",
        help="Konfirmasi bahwa data ujian/hasil akan hilang permanen.",
    )
    parser.add_argument(
        "--clean-mt-archives",
        action="store_true",
        help="Hapus folder arsip MT lokal setelah operasi DB (atau di dry-run hanya laporkan).",
    )
    parser.add_argument(
        "--include-other-published",
        action="store_true",
        help="Juga hapus ujian published yang bukan draft/finished/active (jadwal di masa depan).",
    )
    return parser.parse_args(argv)


def get_database_url() -> str:
    url = os.getenv("DATABASE_URL", "").strip()
    if not url:
        raise CleanupError("DATABASE_URL wajib tersedia di environment.")
    return url


def is_production_url(url: str) -> bool:
    lowered = url.lower()
    return any(pattern in lowered for pattern in PRODUCTION_URL_PATTERNS)


def check_production_safety(url: str, allow_production_write: bool, apply: bool) -> None:
    if not apply:
        return
    if is_production_url(url) and not allow_production_write:
        raise CleanupError(
            "DATABASE_URL terlihat seperti produksi. Gunakan "
            "--allow-production-write bersama --apply dan "
            "--i-understand-destructive."
        )


def normalize_async_database_url(url: str) -> str:
    if url.startswith("postgresql+asyncpg://"):
        return url
    if url.startswith("postgresql://"):
        return "postgresql+asyncpg://" + url[len("postgresql://") :]
    if url.startswith("postgres://"):
        return "postgresql+asyncpg://" + url[len("postgres://") :]
    raise CleanupError("Hanya DATABASE_URL PostgreSQL yang didukung.")


def redacted_database_identity(url: str) -> str:
    safe_url = url.replace("postgresql+asyncpg://", "postgresql://", 1)
    parsed = urlsplit(safe_url)
    host = parsed.hostname or "unknown-host"
    port = f":{parsed.port}" if parsed.port else ""
    database = parsed.path.lstrip("/") or "unknown-database"
    return f"postgresql://{host}{port}/{database}"


def normalize_person_name(value: Any) -> str:
    text = str(value or "").strip().upper()
    text = "".join(
        ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch)
    )
    text = text.replace("’", "'").replace("`", "'")
    text = re.sub(r"[^A-Z0-9' ]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_class_name(value: Any) -> Optional[str]:
    raw = str(value or "").strip().upper()
    if not raw:
        return None
    raw = raw.replace("–", "-").replace("—", "-")
    raw = re.sub(r"\s+", " ", raw).strip()

    compact = re.sub(r"[\s\-]+", "", raw)
    if compact in ROSTER_SHEET_ALIASES:
        return ROSTER_SHEET_ALIASES[compact]

    # X A / XI B / XII C
    match = re.fullmatch(r"(X|XI|XII)[\s\-]*([A-F])", raw)
    if match:
        return f"{match.group(1)} {match.group(2)}"

    # Compact sheet style XA / XIB / XIIA
    match = re.fullmatch(r"(X|XI|XII)([A-F])", compact)
    if match:
        return f"{match.group(1)} {match.group(2)}"

    # Grade-only or major-style XII IPA 1 → keep grade marker only when useful
    match = re.fullmatch(r"(X|XI|XII)(?:[\s\-].*)?", raw)
    if match and re.search(r"(X|XI|XII)", compact):
        # If it has a single trailing letter section somewhere, prefer that.
        section = re.search(r"(X|XI|XII).*?([A-F])\b", raw)
        if section and not re.search(r"\bIPA\b|\bIPS\b|\bIIS\b|\bMIA\b", raw):
            return f"{section.group(1)} {section.group(2)}"
        return match.group(1)
    return None


def class_grade(value: Any) -> Optional[str]:
    canonical = normalize_class_name(value)
    if canonical:
        return canonical.split()[0]
    raw = str(value or "").strip().upper()
    if not raw:
        return None
    compact = re.sub(r"[\s\-]+", "", raw)
    for grade in ("XII", "XI", "X"):
        if compact.startswith(grade):
            return grade
    return None


def is_class_twelve(value: Any) -> bool:
    return class_grade(value) == "XII"


def is_class_ten_or_eleven(value: Any) -> bool:
    return class_grade(value) in {"X", "XI"}


def slugify_username(full_name: str, max_length: int = 40) -> str:
    text = normalize_person_name(full_name).lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    if not text:
        text = "siswa"
    return text[:max_length]


def generate_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits
    # Ensure mix of classes.
    parts = [
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.digits),
    ]
    parts.extend(secrets.choice(alphabet) for _ in range(max(0, length - len(parts))))
    random.SystemRandom().shuffle(parts)
    return "".join(parts)


def hash_password(password: str) -> str:
    try:
        from app.core.security import get_password_hash

        return get_password_hash(password)
    except Exception:
        from passlib.context import CryptContext

        return CryptContext(schemes=["bcrypt"], deprecated="auto").hash(password)


def ensure_private_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    path.chmod(stat.S_IRWXU)
    return path


def write_json(path: Path, payload: Mapping[str, Any]) -> None:
    ensure_private_dir(path.parent)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    path.chmod(stat.S_IRUSR | stat.S_IWUSR)


def write_csv(path: Path, rows: Sequence[Mapping[str, Any]], fieldnames: Sequence[str]) -> None:
    ensure_private_dir(path.parent)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames))
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})
    path.chmod(stat.S_IRUSR | stat.S_IWUSR)


def sheet_to_class_label(sheet_name: str) -> Optional[str]:
    compact = re.sub(r"\s+", "", str(sheet_name or "").strip().upper())
    if any(compact.startswith(prefix) for prefix in IGNORED_SHEET_PREFIXES):
        # XII letter sheets exist but are ignored for import by policy.
        return None
    return ROSTER_SHEET_ALIASES.get(compact) or normalize_class_name(sheet_name)


def load_excel_roster(excel_path: Path) -> list[RosterStudent]:
    if not excel_path.exists():
        raise CleanupError(f"File Excel tidak ditemukan: {excel_path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CleanupError("openpyxl wajib terpasang untuk membaca Excel.") from exc

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    students: list[RosterStudent] = []
    try:
        for sheet_name in workbook.sheetnames:
            class_label = sheet_to_class_label(sheet_name)
            if not class_label or class_grade(class_label) not in {"X", "XI"}:
                continue
            worksheet = workbook[sheet_name]
            rows = list(worksheet.iter_rows(values_only=True))
            header_idx = None
            headers: list[str] = []
            for index, row in enumerate(rows[:12]):
                values = [str(cell).strip().upper() if cell is not None else "" for cell in row]
                if "NAMA SISWA" in values or "NAMA" in values:
                    header_idx = index
                    headers = values
                    break
            if header_idx is None:
                continue

            def find_col(*names: str) -> Optional[int]:
                for name in names:
                    if name in headers:
                        return headers.index(name)
                return None

            col_name = find_col("NAMA SISWA", "NAMA")
            col_nisn = find_col("NISN")
            col_no = find_col("NO")
            if col_name is None:
                continue

            for offset, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
                if not row or col_name >= len(row):
                    continue
                full_name = str(row[col_name] or "").strip()
                if not full_name:
                    continue
                nisn = None
                if col_nisn is not None and col_nisn < len(row) and row[col_nisn] is not None:
                    nisn = str(row[col_nisn]).strip()
                    if nisn.endswith(".0"):
                        nisn = nisn[:-2]
                    if not nisn:
                        nisn = None
                row_number = offset
                if col_no is not None and col_no < len(row) and row[col_no] is not None:
                    with_context = row[col_no]
                    if isinstance(with_context, (int, float)):
                        row_number = int(with_context)
                students.append(
                    RosterStudent(
                        full_name=full_name,
                        student_class=class_label,
                        source_sheet=sheet_name,
                        row_number=row_number,
                        nisn=nisn,
                    )
                )
    finally:
        workbook.close()

    # Guard against accidental empty parse.
    if not students:
        raise CleanupError("Excel tidak berisi siswa kelas X/XI yang bisa dibaca.")
    return students


def classify_exams(rows: Sequence[Mapping[str, Any]], include_other_published: bool = False) -> ExamPlan:
    plan = ExamPlan()
    for row in rows:
        exam_id = int(row["exam_id"])
        lifecycle = str(row.get("lifecycle") or "")
        if lifecycle == "draft":
            plan.draft_ids.append(exam_id)
        elif lifecycle == "finished":
            plan.finished_ids.append(exam_id)
        elif lifecycle == "active":
            plan.active_ids.append(exam_id)
        else:
            plan.other_ids.append(exam_id)
    plan.target_ids = list(plan.draft_ids) + list(plan.finished_ids)
    if include_other_published:
        plan.target_ids.extend(plan.other_ids)
    plan.target_ids = sorted(set(plan.target_ids))
    return plan


def plan_student_sync(
    roster: Sequence[RosterStudent],
    db_students: Sequence[DbStudent],
) -> MatchPlan:
    plan = MatchPlan()
    pool = [student for student in db_students if is_class_ten_or_eleven(student.student_class)]
    by_name: dict[str, list[DbStudent]] = {}
    for student in pool:
        by_name.setdefault(student.normalized_name, []).append(student)

    used_ids: set[int] = set()
    for entry in roster:
        candidates = [
            student
            for student in by_name.get(entry.normalized_name, [])
            if student.id not in used_ids
        ]
        if len(candidates) == 1:
            student = candidates[0]
            used_ids.add(student.id)
            current = student.canonical_class
            target = entry.student_class
            payload = {
                "user_id": student.id,
                "username": student.username,
                "full_name": student.full_name,
                "old_class": student.student_class,
                "new_class": target,
                "source_sheet": entry.source_sheet,
            }
            if current == target and student.is_active:
                plan.unchanged.append(payload)
            else:
                plan.matched_updates.append(payload)
        elif len(candidates) == 0:
            plan.create_new.append(
                {
                    "full_name": entry.full_name,
                    "student_class": entry.student_class,
                    "source_sheet": entry.source_sheet,
                    "row_number": entry.row_number,
                    "nisn": entry.nisn,
                }
            )
        else:
            plan.ambiguous.append(
                {
                    "full_name": entry.full_name,
                    "student_class": entry.student_class,
                    "source_sheet": entry.source_sheet,
                    "candidate_ids": [item.id for item in candidates],
                    "candidate_usernames": [item.username for item in candidates],
                }
            )

    for student in pool:
        if student.id in used_ids:
            continue
        # Names that were ambiguous stay active for manual handling.
        if any(student.id in item.get("candidate_ids", []) for item in plan.ambiguous):
            continue
        if student.is_active:
            plan.deactivate.append(
                {
                    "user_id": student.id,
                    "username": student.username,
                    "full_name": student.full_name,
                    "student_class": student.student_class,
                }
            )
    return plan


def unique_usernames(full_names: Sequence[str], existing: set[str]) -> list[str]:
    taken = {name.lower() for name in existing}
    result: list[str] = []
    for full_name in full_names:
        base = slugify_username(full_name) or "siswa"
        candidate = base
        suffix = 1
        while candidate.lower() in taken:
            suffix += 1
            candidate = f"{base}{suffix}"[:100]
        taken.add(candidate.lower())
        result.append(candidate)
    return result


def exam_delete_statements(exam_ids: Sequence[int]) -> list[tuple[str, str]]:
    """Return ordered (label, sql) hard-delete statements for exams."""
    if not exam_ids:
        return []
    # selected_option_id has ON DELETE NO ACTION and may lack a supporting index.
    # Null it before deleting options so PostgreSQL does not seq-scan answers.
    return [
        (
            "security_events_by_session",
            """
            DELETE FROM security_events
            WHERE session_id IN (
                SELECT id FROM exam_sessions WHERE exam_id = ANY(:exam_ids)
            )
            """,
        ),
        (
            "exam_logs",
            """
            DELETE FROM exam_logs
            WHERE session_id IN (
                SELECT id FROM exam_sessions WHERE exam_id = ANY(:exam_ids)
            )
            """,
        ),
        (
            "answers",
            """
            DELETE FROM answers
            WHERE session_id IN (
                SELECT id FROM exam_sessions WHERE exam_id = ANY(:exam_ids)
            )
            """,
        ),
        (
            "answers_by_question",
            """
            DELETE FROM answers
            WHERE question_id IN (
                SELECT id FROM questions WHERE exam_id = ANY(:exam_ids)
            )
            """,
        ),
        (
            "null_selected_option_id",
            """
            UPDATE answers
            SET selected_option_id = NULL
            WHERE selected_option_id IN (
                SELECT qo.id
                FROM question_options qo
                JOIN questions q ON q.id = qo.question_id
                WHERE q.exam_id = ANY(:exam_ids)
            )
            """,
        ),
        (
            "exam_sessions",
            "DELETE FROM exam_sessions WHERE exam_id = ANY(:exam_ids)",
        ),
        (
            "question_tags_map",
            """
            DELETE FROM question_tags_map
            WHERE question_id IN (
                SELECT id FROM questions WHERE exam_id = ANY(:exam_ids)
            )
            """,
        ),
        (
            "question_options",
            """
            DELETE FROM question_options
            WHERE question_id IN (
                SELECT id FROM questions WHERE exam_id = ANY(:exam_ids)
            )
            """,
        ),
        (
            "questions",
            "DELETE FROM questions WHERE exam_id = ANY(:exam_ids)",
        ),
        (
            "scheduled_publications",
            "DELETE FROM scheduled_publications WHERE exam_id = ANY(:exam_ids)",
        ),
        (
            "exams",
            "DELETE FROM exams WHERE id = ANY(:exam_ids)",
        ),
    ]


def user_predelete_statements() -> list[tuple[str, str]]:
    return [
        (
            "security_events_by_user",
            "DELETE FROM security_events WHERE user_id = ANY(:user_ids)",
        ),
        (
            "refresh_tokens",
            "DELETE FROM refresh_tokens WHERE user_id = ANY(:user_ids)",
        ),
        (
            "notifications",
            "DELETE FROM notifications WHERE user_id = ANY(:user_ids)",
        ),
        (
            "user_activity_logs",
            "DELETE FROM user_activity_logs WHERE user_id = ANY(:user_ids)",
        ),
        (
            "security_events_by_remaining_sessions",
            """
            DELETE FROM security_events
            WHERE session_id IN (
                SELECT id FROM exam_sessions WHERE user_id = ANY(:user_ids)
            )
            """,
        ),
        (
            "exam_logs_by_user_sessions",
            """
            DELETE FROM exam_logs
            WHERE session_id IN (
                SELECT id FROM exam_sessions WHERE user_id = ANY(:user_ids)
            )
            """,
        ),
        (
            "answers_by_user_sessions",
            """
            DELETE FROM answers
            WHERE session_id IN (
                SELECT id FROM exam_sessions WHERE user_id = ANY(:user_ids)
            )
            """,
        ),
        (
            "exam_sessions_by_user",
            "DELETE FROM exam_sessions WHERE user_id = ANY(:user_ids)",
        ),
    ]


async def execute_named(conn: Any, sql: str, **params: Any) -> str:
    """Execute SQL with :name params using asyncpg $n placeholders."""
    names: list[str] = []

    def replacer(match: re.Match[str]) -> str:
        name = match.group(1)
        if name not in names:
            names.append(name)
        return f"${names.index(name) + 1}"

    converted = re.sub(r":([a-zA-Z_][a-zA-Z0-9_]*)", replacer, sql)
    values = [params[name] for name in names]
    return await conn.execute(converted, *values)


async def fetch_named(conn: Any, sql: str, **params: Any) -> list[Any]:
    names: list[str] = []

    def replacer(match: re.Match[str]) -> str:
        name = match.group(1)
        if name not in names:
            names.append(name)
        return f"${names.index(name) + 1}"

    converted = re.sub(r":([a-zA-Z_][a-zA-Z0-9_]*)", replacer, sql)
    values = [params[name] for name in names]
    return await conn.fetch(converted, *values)


async def fetchrow_named(conn: Any, sql: str, **params: Any) -> Any:
    rows = await fetch_named(conn, sql, **params)
    return rows[0] if rows else None


async def list_public_tables(conn: Any) -> set[str]:
    rows = await conn.fetch(
        """
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'public'
        """
    )
    return {str(row["table_name"]) for row in rows}


def chunked(values: Sequence[int], size: int) -> list[list[int]]:
    items = [int(value) for value in values]
    if size <= 0:
        return [items]
    return [items[index : index + size] for index in range(0, len(items), size)]


async def hard_delete_exams(
    conn: Any,
    exam_ids: Sequence[int],
    tables: set[str],
    batch_size: int = 20,
) -> list[dict[str, Any]]:
    """Hard-delete exams in small committed batches to avoid long locks/timeouts."""
    results: list[dict[str, Any]] = []
    if not exam_ids:
        return results

    primary_by_label = {
        "security_events_by_session": "security_events",
        "exam_logs": "exam_logs",
        "answers": "answers",
        "answers_by_question": "answers",
        "null_selected_option_id": "answers",
        "exam_sessions": "exam_sessions",
        "question_tags_map": "question_tags_map",
        "question_options": "question_options",
        "questions": "questions",
        "scheduled_publications": "scheduled_publications",
        "exams": "exams",
    }

    # Prefer short statements and autocommit-like batches.
    await conn.execute("SET statement_timeout = '0'")
    await conn.execute("SET lock_timeout = '0'")

    for batch_index, ids in enumerate(chunked(exam_ids, batch_size), start=1):
        batch_result: dict[str, Any] = {"batch": batch_index, "exam_ids": ids, "steps": []}
        async with conn.transaction():
            for label, sql in exam_delete_statements(ids):
                primary = primary_by_label[label]
                if primary not in tables:
                    batch_result["steps"].append({"label": label, "skipped": True})
                    continue
                status = await execute_named(conn, sql, exam_ids=ids)
                batch_result["steps"].append({"label": label, "status": status})
        results.append(batch_result)
        # Progress to stdout so long runs are visible.
        print(
            json.dumps(
                {
                    "progress": "exam_batch_deleted",
                    "batch": batch_index,
                    "batch_size": len(ids),
                    "exam_ids_sample": ids[:5],
                },
                ensure_ascii=False,
            ),
            flush=True,
        )
    return results


async def hard_delete_users(
    conn: Any,
    user_ids: Sequence[int],
    tables: set[str],
    batch_size: int = 50,
) -> dict[str, Any]:
    deleted: list[int] = []
    failed: list[dict[str, Any]] = []
    deactivated: list[int] = []
    if not user_ids:
        return {"deleted": deleted, "failed": failed, "deactivated": deactivated}

    primary_by_label = {
        "security_events_by_user": "security_events",
        "refresh_tokens": "refresh_tokens",
        "notifications": "notifications",
        "user_activity_logs": "user_activity_logs",
        "security_events_by_remaining_sessions": "security_events",
        "exam_logs_by_user_sessions": "exam_logs",
        "answers_by_user_sessions": "answers",
        "exam_sessions_by_user": "exam_sessions",
    }

    for batch in chunked(user_ids, batch_size):
        # Shared child cleanup for this batch first.
        async with conn.transaction():
            for label, sql in user_predelete_statements():
                primary = primary_by_label[label]
                if primary not in tables:
                    continue
                await execute_named(conn, sql, user_ids=batch)

        for user_id in batch:
            try:
                async with conn.transaction():
                    status = await execute_named(
                        conn,
                        "DELETE FROM users WHERE id = :user_id AND lower(trim(role)) = 'student'",
                        user_id=user_id,
                    )
                    if status.endswith(" 0"):
                        raise CleanupError(f"user {user_id} not deleted")
                deleted.append(user_id)
            except Exception as exc:  # noqa: BLE001 - per-user fallback is intentional
                try:
                    async with conn.transaction():
                        await execute_named(
                            conn,
                            """
                            UPDATE users
                            SET is_active = FALSE
                            WHERE id = :user_id AND lower(trim(role)) = 'student'
                            """,
                            user_id=user_id,
                        )
                    deactivated.append(user_id)
                    failed.append(
                        {
                            "user_id": user_id,
                            "error": str(exc),
                            "fallback": "deactivated",
                        }
                    )
                except Exception as deactivate_exc:  # noqa: BLE001
                    failed.append(
                        {
                            "user_id": user_id,
                            "error": str(exc),
                            "deactivate_error": str(deactivate_exc),
                            "fallback": "none",
                        }
                    )
        print(
            json.dumps(
                {
                    "progress": "class12_batch",
                    "deleted_so_far": len(deleted),
                    "deactivated_so_far": len(deactivated),
                    "failed_so_far": len(failed),
                },
                ensure_ascii=False,
            ),
            flush=True,
        )
    return {"deleted": deleted, "failed": failed, "deactivated": deactivated}


def mt_archive_targets(repo_root: Path) -> list[Path]:
    return [(repo_root / relative).resolve() for relative in DEFAULT_MT_ARCHIVES]


def clean_mt_archives(repo_root: Path, apply: bool) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    for path in mt_archive_targets(repo_root):
        exists = path.exists()
        entry = {
            "path": str(path),
            "exists": exists,
            "action": "would_delete" if exists and not apply else ("deleted" if exists and apply else "missing"),
        }
        if exists and apply:
            shutil.rmtree(path)
            entry["action"] = "deleted"
        results.append(entry)
    return results


async def collect_snapshot(
    conn: Any,
    excel_path: Path,
    include_other_published: bool,
) -> dict[str, Any]:
    captured_at = await conn.fetchval("SELECT NOW()")
    exam_rows = await fetch_named(conn, EXAM_TARGET_SQL, captured_at=captured_at)
    exam_plan = classify_exams(exam_rows, include_other_published=include_other_published)

    child_counts = {
        "session_count": 0,
        "answer_count": 0,
        "log_count": 0,
        "question_count": 0,
    }
    if exam_plan.target_ids:
        # Optional tables may not exist in all environments.
        tables = await list_public_tables(conn)
        session_count = await conn.fetchval(
            "SELECT COUNT(*) FROM exam_sessions WHERE exam_id = ANY($1::int[])",
            exam_plan.target_ids,
        )
        child_counts["session_count"] = int(session_count or 0)
        if "answers" in tables:
            answer_count = await conn.fetchval(
                """
                SELECT COUNT(*) FROM answers
                WHERE session_id IN (
                    SELECT id FROM exam_sessions WHERE exam_id = ANY($1::int[])
                )
                """,
                exam_plan.target_ids,
            )
            child_counts["answer_count"] = int(answer_count or 0)
        if "exam_logs" in tables:
            log_count = await conn.fetchval(
                """
                SELECT COUNT(*) FROM exam_logs
                WHERE session_id IN (
                    SELECT id FROM exam_sessions WHERE exam_id = ANY($1::int[])
                )
                """,
                exam_plan.target_ids,
            )
            child_counts["log_count"] = int(log_count or 0)
        question_count = await conn.fetchval(
            "SELECT COUNT(*) FROM questions WHERE exam_id = ANY($1::int[])",
            exam_plan.target_ids,
        )
        child_counts["question_count"] = int(question_count or 0)
    exam_plan.child_counts = child_counts

    student_rows = await conn.fetch(STUDENT_SQL)
    db_students = [
        DbStudent(
            id=int(row["id"]),
            username=str(row["username"]),
            full_name=str(row["full_name"]),
            student_class=row["student_class"],
            is_active=bool(row["is_active"]),
        )
        for row in student_rows
    ]
    class12 = [student for student in db_students if is_class_twelve(student.student_class)]
    roster = load_excel_roster(excel_path)
    match_plan = plan_student_sync(roster, db_students)

    return {
        "captured_at": captured_at,
        "exam_plan": exam_plan,
        "db_students": db_students,
        "class12": class12,
        "roster": roster,
        "match_plan": match_plan,
        "all_usernames": {str(row["username"]) for row in await conn.fetch("SELECT username FROM users")},
    }


async def apply_changes(
    conn: Any,
    snapshot: Mapping[str, Any],
) -> dict[str, Any]:
    tables = await list_public_tables(conn)
    exam_plan: ExamPlan = snapshot["exam_plan"]
    class12: list[DbStudent] = snapshot["class12"]
    match_plan: MatchPlan = snapshot["match_plan"]
    all_usernames: set[str] = set(snapshot["all_usernames"])

    print(
        json.dumps(
            {
                "progress": "start_exam_purge",
                "target_exams": len(exam_plan.target_ids),
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    exam_result = await hard_delete_exams(conn, exam_plan.target_ids, tables, batch_size=5)

    print(
        json.dumps(
            {
                "progress": "start_class12_delete",
                "target_users": len(class12),
            },
            ensure_ascii=False,
        ),
        flush=True,
    )
    class12_result = await hard_delete_users(
        conn,
        [item.id for item in class12],
        tables,
        batch_size=40,
    )

    update_count = 0
    for batch in chunked([int(item["user_id"]) for item in match_plan.matched_updates], 100):
        # Keep class mapping from original plan.
        by_id = {int(item["user_id"]): item for item in match_plan.matched_updates}
        async with conn.transaction():
            for user_id in batch:
                item = by_id[user_id]
                await execute_named(
                    conn,
                    """
                    UPDATE users
                    SET student_class = :student_class,
                        is_active = TRUE
                    WHERE id = :user_id
                      AND lower(trim(role)) = 'student'
                    """,
                    student_class=item["new_class"],
                    user_id=user_id,
                )
                update_count += 1
    print(
        json.dumps({"progress": "classes_updated", "count": update_count}, ensure_ascii=False),
        flush=True,
    )

    deactivate_count = 0
    for batch in chunked([int(item["user_id"]) for item in match_plan.deactivate], 100):
        async with conn.transaction():
            for user_id in batch:
                await execute_named(
                    conn,
                    """
                    UPDATE users
                    SET is_active = FALSE
                    WHERE id = :user_id
                      AND lower(trim(role)) = 'student'
                    """,
                    user_id=user_id,
                )
                deactivate_count += 1
    print(
        json.dumps(
            {"progress": "old_students_deactivated", "count": deactivate_count},
            ensure_ascii=False,
        ),
        flush=True,
    )

    created_rows: list[dict[str, Any]] = []
    if match_plan.create_new:
        usernames = unique_usernames(
            [item["full_name"] for item in match_plan.create_new],
            all_usernames,
        )
        for item, username in zip(match_plan.create_new, usernames):
            password = generate_password()
            password_hash = hash_password(password)
            async with conn.transaction():
                row = await fetchrow_named(
                    conn,
                    """
                    INSERT INTO users (
                        username, password_hash, full_name, role, student_class, is_active, created_at
                    ) VALUES (
                        :username, :password_hash, :full_name, 'student', :student_class, TRUE, NOW()
                    )
                    RETURNING id, username
                    """,
                    username=username,
                    password_hash=password_hash,
                    full_name=item["full_name"],
                    student_class=item["student_class"],
                )
            created_rows.append(
                {
                    "user_id": int(row["id"]),
                    "username": row["username"],
                    "password": password,
                    "full_name": item["full_name"],
                    "student_class": item["student_class"],
                    "source_sheet": item["source_sheet"],
                }
            )
            all_usernames.add(username)
        print(
            json.dumps(
                {"progress": "students_created", "count": len(created_rows)},
                ensure_ascii=False,
            ),
            flush=True,
        )

    return {
        "exam_delete": exam_result,
        "class12": class12_result,
        "updated_classes": update_count,
        "deactivated": deactivate_count,
        "created_students": [
            {key: value for key, value in row.items() if key != "password"}
            for row in created_rows
        ],
        "created_credentials": created_rows,
        "ambiguous_left": match_plan.ambiguous,
    }


def build_report(
    *,
    mode: str,
    database_identity: str,
    excel_path: Path,
    snapshot: Mapping[str, Any],
    apply_result: Optional[Mapping[str, Any]] = None,
    mt_cleanup: Optional[Sequence[Mapping[str, Any]]] = None,
) -> dict[str, Any]:
    exam_plan: ExamPlan = snapshot["exam_plan"]
    match_plan: MatchPlan = snapshot["match_plan"]
    class12: list[DbStudent] = snapshot["class12"]
    roster: list[RosterStudent] = snapshot["roster"]
    return {
        "mode": mode,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "database": database_identity,
        "excel": str(excel_path),
        "exams": {
            "draft": len(exam_plan.draft_ids),
            "finished": len(exam_plan.finished_ids),
            "active_preserved": len(exam_plan.active_ids),
            "other_published": len(exam_plan.other_ids),
            "target_delete": len(exam_plan.target_ids),
            "target_ids_sample": exam_plan.target_ids[:50],
            "child_counts": exam_plan.child_counts,
        },
        "class12": {
            "count": len(class12),
            "sample": [
                {
                    "id": item.id,
                    "username": item.username,
                    "full_name": item.full_name,
                    "student_class": item.student_class,
                }
                for item in class12[:50]
            ],
        },
        "roster_x_xi": {
            "count": len(roster),
            "by_class": _count_by(roster, key=lambda item: item.student_class),
        },
        "sync_plan": {
            "matched_updates": len(match_plan.matched_updates),
            "create_new": len(match_plan.create_new),
            "deactivate": len(match_plan.deactivate),
            "unchanged": len(match_plan.unchanged),
            "ambiguous": len(match_plan.ambiguous),
            "ambiguous_details": match_plan.ambiguous,
            "create_new_sample": match_plan.create_new[:50],
            "deactivate_sample": match_plan.deactivate[:50],
            "update_sample": match_plan.matched_updates[:50],
        },
        "apply_result": apply_result,
        "mt_cleanup": list(mt_cleanup or []),
        "notes": [
            "NISN Excel tidak dipakai sebagai identitas unik.",
            "Ujian aktif (published & masih berjalan) tidak dihapus.",
            "Sheet XII di Excel tidak diimpor ulang.",
            "Hard delete bersifat permanen; tidak ada rollback DB.",
        ],
    }


def _count_by(items: Sequence[Any], key) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in items:
        label = str(key(item))
        counts[label] = counts.get(label, 0) + 1
    return dict(sorted(counts.items()))


async def async_main(args: argparse.Namespace) -> int:
    apply = bool(args.apply)
    if apply and not args.i_understand_destructive:
        raise CleanupError(
            "Mode --apply membutuhkan --i-understand-destructive."
        )
    if apply and not args.allow_production_write:
        # Still allow non-prod without the production flag, but check URL below.
        pass

    database_url = get_database_url()
    check_production_safety(database_url, args.allow_production_write, apply)
    if apply and is_production_url(database_url) and not args.i_understand_destructive:
        raise CleanupError("Produksi write membutuhkan --i-understand-destructive.")

    async_url = normalize_async_database_url(database_url)
    # asyncpg accepts postgresql:// (not SQLAlchemy's +asyncpg driver suffix).
    asyncpg_url = (
        async_url.replace("postgresql+asyncpg://", "postgresql://", 1)
        .replace("postgres+asyncpg://", "postgres://", 1)
    )
    database_identity = redacted_database_identity(database_url)
    excel_path = args.excel.expanduser().resolve()
    report_dir = ensure_private_dir(args.report_dir.expanduser().resolve())
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    mode = "apply" if apply else "dry-run"

    try:
        import asyncpg
    except ImportError as exc:
        raise CleanupError("asyncpg wajib terpasang.") from exc

    conn = await asyncpg.connect(asyncpg_url)
    apply_result = None
    try:
        # Snapshot first (read). Apply uses many small committed batches so a
        # timeout cannot roll back an entire multi-hour purge.
        snapshot = await collect_snapshot(
            conn,
            excel_path,
            include_other_published=bool(args.include_other_published),
        )
        if apply:
            apply_result = await apply_changes(conn, snapshot)
    finally:
        await conn.close()

    repo_root = Path(__file__).resolve().parents[1]
    mt_cleanup = None
    if args.clean_mt_archives:
        mt_cleanup = clean_mt_archives(repo_root, apply=apply)

    report = build_report(
        mode=mode,
        database_identity=database_identity,
        excel_path=excel_path,
        snapshot=snapshot,
        apply_result=(
            None
            if apply_result is None
            else {
                **{k: v for k, v in apply_result.items() if k != "created_credentials"},
            }
        ),
        mt_cleanup=mt_cleanup,
    )
    report_path = report_dir / f"year_rollover_{mode}_{stamp}.json"
    write_json(report_path, report)

    credentials_path = None
    if apply_result and apply_result.get("created_credentials"):
        credentials_path = report_dir / f"new_students_credentials_{stamp}.csv"
        write_csv(
            credentials_path,
            apply_result["created_credentials"],
            fieldnames=[
                "user_id",
                "username",
                "password",
                "full_name",
                "student_class",
                "source_sheet",
            ],
        )

    print(
        json.dumps(
            {
                "mode": mode,
                "database": database_identity,
                "report": str(report_path),
                "credentials": str(credentials_path) if credentials_path else None,
                "exams_target": report["exams"]["target_delete"],
                "class12": report["class12"]["count"],
                "sync": {
                    "update": report["sync_plan"]["matched_updates"],
                    "create": report["sync_plan"]["create_new"],
                    "deactivate": report["sync_plan"]["deactivate"],
                    "ambiguous": report["sync_plan"]["ambiguous"],
                },
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        return asyncio.run(async_main(args))
    except CleanupError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
