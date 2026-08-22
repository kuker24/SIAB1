#!/usr/bin/env python3
"""Create a read-only maintenance archive for completed teacher exams.

The database is never modified. The only writes are PDF/XLSX/JSON files under
an operator-selected local output directory.
"""
from __future__ import annotations

import argparse
import asyncio
import hashlib
import html
import importlib.metadata
import json
import os
import re
import shutil
import stat
import sys
from collections import defaultdict
from contextlib import suppress
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional, Sequence
from urllib.parse import urlsplit

TERMINAL_STATUSES = ("submitted", "completed")
PRODUCTION_URL_PATTERNS = ("103.175.218.56", "man1rokanhulu.cloud", "adminujian")
SAFE_SETTINGS_KEYS = (
    "acceptable_answers",
    "answer_key",
    "correct_answer",
    "correct_statements",
    "model_answer",
    "statement_answers",
    "statements",
)
REQUIRED_QUESTION_SHEETS = (
    "Info_Ujian",
    "Soal",
    "Pilihan",
    "Kunci_Jawaban",
    "Referensi_Media",
)
REQUIRED_RESULT_SHEETS = (
    "Ringkasan",
    "Hasil_Terbaru",
    "Semua_Percobaan",
    "Metadata_Ekspor",
)

EXAM_SELECTION_SQL = """
SELECT
    u.id AS teacher_id,
    u.username AS teacher_username,
    u.full_name AS teacher_name,
    u.role AS teacher_role,
    u.is_active AS teacher_is_active,
    e.id AS exam_id,
    e.title AS exam_title,
    e.description AS exam_description,
    e.subject,
    e.exam_type,
    e.academic_year,
    e.duration_minutes,
    e.start_time,
    e.end_time,
    e.passing_score,
    e.is_published,
    e.is_deleted,
    e.deleted_at,
    e.created_at,
    e.updated_at,
    COUNT(q.id) AS question_count
FROM users u
JOIN exams e ON e.creator_id = u.id
JOIN questions q ON q.exam_id = e.id
WHERE (
    lower(trim(u.role)) IN ('teacher', 'guruplus')
    OR lower(trim(u.username)) = 'kamad'
  )
  AND e.is_published IS TRUE
  AND e.end_time IS NOT NULL
  AND e.end_time <= :captured_at
GROUP BY
    u.id, u.username, u.full_name, u.role, u.is_active,
    e.id, e.title, e.description, e.subject, e.exam_type,
    e.academic_year, e.duration_minutes, e.start_time, e.end_time,
    e.passing_score, e.is_published, e.is_deleted, e.deleted_at,
    e.created_at, e.updated_at
ORDER BY lower(u.full_name), u.id, e.end_time, e.id
"""

QUESTION_SQL = """
SELECT
    q.id AS question_id,
    q.question_text,
    q.stimulus,
    q.question_type,
    q.question_subtype,
    q.pgk_type,
    q.difficulty_level,
    q.question_settings,
    q.points,
    q.order_index,
    q.image_url,
    q.video_url,
    q.audio_url
FROM questions q
WHERE q.exam_id = :exam_id
ORDER BY q.order_index, q.id
"""

OPTION_SQL = """
SELECT
    qo.id AS option_id,
    qo.question_id,
    qo.option_text,
    qo.is_correct,
    qo.order_index,
    qo.option_group,
    qo.pair_id,
    qo.option_metadata
FROM question_options qo
JOIN questions q ON q.id = qo.question_id
WHERE q.exam_id = :exam_id
ORDER BY q.order_index, q.id, qo.order_index, qo.id
"""

SESSION_SQL = """
SELECT
    es.id AS session_id,
    es.user_id,
    u.full_name AS student_name,
    u.student_class,
    es.start_time,
    es.end_time,
    es.status,
    es.score,
    es.violation_count,
    es.total_paused_seconds
FROM exam_sessions es
JOIN users u ON u.id = es.user_id
WHERE es.exam_id = :exam_id
  AND es.status IN ('submitted', 'completed')
ORDER BY es.end_time DESC NULLS LAST, es.id DESC
"""


class ArchiveError(RuntimeError):
    """Raised when the archive cannot be created safely."""


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("MT/MT2026-08-10"),
        help="Final archive directory (must not already exist).",
    )
    parser.add_argument(
        "--allow-production-readonly",
        action="store_true",
        help="Allow an explicitly read-only run against production.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Count selected data without writing archive files.",
    )
    return parser.parse_args(argv)


def get_database_url() -> str:
    url = os.getenv("DATABASE_URL", "").strip()
    if not url:
        raise ArchiveError("DATABASE_URL wajib tersedia di environment.")
    return url


def is_production_url(url: str) -> bool:
    lowered = url.lower()
    return any(pattern in lowered for pattern in PRODUCTION_URL_PATTERNS)


def check_production_safety(url: str, allow_production: bool) -> None:
    if is_production_url(url) and not allow_production:
        raise ArchiveError(
            "DATABASE_URL terlihat seperti produksi. Gunakan "
            "--allow-production-readonly untuk proses hanya-baca."
        )


def normalize_async_database_url(url: str) -> str:
    if url.startswith("postgresql+asyncpg://"):
        return url
    if url.startswith("postgresql://"):
        return "postgresql+asyncpg://" + url[len("postgresql://") :]
    if url.startswith("postgres://"):
        return "postgresql+asyncpg://" + url[len("postgres://") :]
    raise ArchiveError("Hanya DATABASE_URL PostgreSQL yang didukung.")


def redacted_database_identity(url: str) -> str:
    safe_url = url.replace("postgresql+asyncpg://", "postgresql://", 1)
    parsed = urlsplit(safe_url)
    host = parsed.hostname or "unknown-host"
    port = f":{parsed.port}" if parsed.port else ""
    database = parsed.path.lstrip("/") or "unknown-database"
    return f"postgresql://{host}{port}/{database}"


def slugify(value: Any, fallback: str = "tanpa_nama", max_length: int = 80) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text, flags=re.ASCII)
    text = re.sub(r"_+", "_", text).strip("_")
    return (text or fallback)[:max_length].rstrip("_") or fallback


def json_value(value: Any) -> Any:
    if isinstance(value, Mapping):
        return {str(key): json_value(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_value(item) for item in value]
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def safe_settings(settings: Any) -> dict[str, Any]:
    if not isinstance(settings, Mapping):
        return {}
    return {
        key: json_value(settings[key])
        for key in SAFE_SETTINGS_KEYS
        if key in settings
    }


def answer_key(question: Mapping[str, Any], options: Sequence[Mapping[str, Any]]) -> str:
    correct_options = [
        str(option.get("option_text") or "")
        for option in options
        if option.get("is_correct") is True
    ]
    parts: list[str] = []
    if correct_options:
        parts.append("Pilihan benar: " + " | ".join(correct_options))

    selected_settings = safe_settings(question.get("question_settings"))
    for key, value in selected_settings.items():
        rendered = json.dumps(json_value(value), ensure_ascii=False, sort_keys=True)
        parts.append(f"{key}: {rendered}")
    return "\n".join(parts) or "Kunci belum tersedia"


def pick_latest_scored_sessions(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    """Reuse the result-view rule: input rows must be newest first."""
    latest_any: dict[int, dict[str, Any]] = {}
    latest_scored: dict[int, dict[str, Any]] = {}
    for original in rows:
        row = dict(original)
        user_id = int(row["user_id"])
        latest_any.setdefault(user_id, row)
        if row.get("score") is not None:
            latest_scored.setdefault(user_id, row)

    selected = [latest_scored.get(user_id, row) for user_id, row in latest_any.items()]
    minimum = datetime.min.replace(tzinfo=timezone.utc)
    selected.sort(
        key=lambda row: (
            row.get("end_time") is not None,
            _aware_datetime(row.get("end_time")) or minimum,
            int(row.get("session_id") or 0),
        ),
        reverse=True,
    )
    return selected


def result_summary(rows: Sequence[Mapping[str, Any]], passing_score: float) -> dict[str, Any]:
    scored = [float(row["score"]) for row in rows if row.get("score") is not None]
    passed = sum(score >= passing_score for score in scored)
    return {
        "participants": len(rows),
        "scored": len(scored),
        "ungraded": len(rows) - len(scored),
        "average": sum(scored) / len(scored) if scored else None,
        "highest": max(scored) if scored else None,
        "lowest": min(scored) if scored else None,
        "passed": passed,
        "failed": len(scored) - passed,
        "pass_rate": (passed / len(scored) * 100) if scored else None,
    }


def _aware_datetime(value: Any) -> Optional[datetime]:
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def duration_seconds(row: Mapping[str, Any]) -> Optional[int]:
    start = _aware_datetime(row.get("start_time"))
    end = _aware_datetime(row.get("end_time"))
    if not start or not end:
        return None
    return max(0, int((end - start).total_seconds()))


def format_datetime(value: Any) -> str:
    parsed = _aware_datetime(value)
    return parsed.isoformat() if parsed else ""


def format_score(value: Any) -> str:
    return "Belum dinilai" if value is None else f"{float(value):.2f}"


def _clean_pdf_text(value: Any) -> str:
    plain = re.sub(r"</?[A-Za-z][^>]*>", "", str(value or ""))
    plain = "".join(char for char in plain if char in "\n\t" or ord(char) >= 32)
    return html.escape(plain.encode("latin-1", "replace").decode("latin-1"))


def _ensure_export_dependencies() -> None:
    missing: list[str] = []
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")
    try:
        import reportlab  # noqa: F401
    except ImportError:
        missing.append("reportlab")
    if missing:
        raise ArchiveError(
            "Dependency ekspor belum tersedia: " + ", ".join(missing) + "."
        )


def _secure_mkdir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=False, mode=0o700)
    path.chmod(0o700)


def _write_bytes(path: Path, content: bytes) -> None:
    path.write_bytes(content)
    path.chmod(0o600)


def _write_json(path: Path, payload: Any) -> None:
    _write_bytes(
        path,
        (json.dumps(json_value(payload), ensure_ascii=False, indent=2) + "\n").encode("utf-8"),
    )


def _style_workbook(workbook: Any) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="1F2937")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            width = max((len(str(cell.value or "")) for cell in column_cells), default=0)
            sheet.column_dimensions[letter].width = min(max(width + 2, 10), 60)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)


def create_question_xlsx(
    exam: Mapping[str, Any],
    questions: Sequence[Mapping[str, Any]],
    options: Sequence[Mapping[str, Any]],
) -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    info = workbook.active
    if info is None:
        raise ArchiveError("Workbook soal gagal dibuat.")
    info.title = "Info_Ujian"
    info.append(["Field", "Nilai"])
    for label, key in (
        ("ID Ujian", "exam_id"),
        ("Judul", "exam_title"),
        ("Mata Pelajaran", "subject"),
        ("Jenis Ujian", "exam_type"),
        ("Tahun Ajaran", "academic_year"),
        ("Guru", "teacher_name"),
        ("Mulai", "start_time"),
        ("Selesai", "end_time"),
        ("Durasi (menit)", "duration_minutes"),
        ("Jumlah Soal", "question_count"),
        ("Soft-delete", "is_deleted"),
    ):
        value = exam.get(key)
        info.append([label, format_datetime(value) if isinstance(value, datetime) else json_value(value)])

    question_sheet = workbook.create_sheet("Soal")
    question_sheet.append([
        "No", "ID Soal", "Urutan", "Tipe", "Subtipe", "Tipe PGK",
        "Kesulitan", "Poin", "Stimulus", "Teks Soal",
    ])
    for number, question in enumerate(questions, 1):
        question_sheet.append([
            number,
            question.get("question_id"),
            question.get("order_index"),
            question.get("question_type"),
            question.get("question_subtype"),
            question.get("pgk_type"),
            question.get("difficulty_level"),
            float(question.get("points") or 0),
            question.get("stimulus") or "",
            question.get("question_text") or "",
        ])

    option_sheet = workbook.create_sheet("Pilihan")
    option_sheet.append([
        "ID Soal", "ID Pilihan", "Urutan", "Teks Pilihan", "Benar",
        "Grup", "Pair ID", "Metadata",
    ])
    for option in options:
        option_sheet.append([
            option.get("question_id"),
            option.get("option_id"),
            option.get("order_index"),
            option.get("option_text") or "",
            "Ya" if option.get("is_correct") is True else "Tidak",
            option.get("option_group") or "",
            option.get("pair_id") or "",
            json.dumps(json_value(option.get("option_metadata") or {}), ensure_ascii=False, sort_keys=True),
        ])

    options_by_question: dict[int, list[Mapping[str, Any]]] = defaultdict(list)
    for option in options:
        options_by_question[int(option["question_id"])].append(option)

    key_sheet = workbook.create_sheet("Kunci_Jawaban")
    key_sheet.append(["ID Soal", "Urutan", "Tipe", "Kunci Jawaban"])
    for question in questions:
        question_id = int(question["question_id"])
        key_sheet.append([
            question_id,
            question.get("order_index"),
            question.get("question_type"),
            answer_key(question, options_by_question[question_id]),
        ])

    media_sheet = workbook.create_sheet("Referensi_Media")
    media_sheet.append(["ID Soal", "Jenis Media", "URL atau Path"])
    for question in questions:
        for key, label in (("image_url", "Gambar"), ("video_url", "Video"), ("audio_url", "Audio")):
            if question.get(key):
                media_sheet.append([question.get("question_id"), label, question[key]])

    _style_workbook(workbook)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_results_xlsx(
    exam: Mapping[str, Any],
    latest: Sequence[Mapping[str, Any]],
    all_attempts: Sequence[Mapping[str, Any]],
    captured_at: datetime,
) -> bytes:
    from openpyxl import Workbook

    passing_score = float(exam.get("passing_score") or 70)
    summary = result_summary(latest, passing_score)
    workbook = Workbook()
    summary_sheet = workbook.active
    if summary_sheet is None:
        raise ArchiveError("Workbook hasil gagal dibuat.")
    summary_sheet.title = "Ringkasan"
    summary_sheet.append(["Statistik", "Nilai"])
    summary_sheet.append(["Judul Ujian", exam.get("exam_title")])
    summary_sheet.append(["Guru", exam.get("teacher_name")])
    summary_sheet.append(["KKM", passing_score])
    for key, label in (
        ("participants", "Peserta resmi"),
        ("scored", "Sudah dinilai"),
        ("ungraded", "Belum dinilai"),
        ("average", "Rata-rata"),
        ("highest", "Tertinggi"),
        ("lowest", "Terendah"),
        ("passed", "Lulus"),
        ("failed", "Tidak lulus"),
        ("pass_rate", "Persentase lulus"),
    ):
        summary_sheet.append([label, summary[key]])

    headers = [
        "ID Siswa", "Nama", "Kelas", "ID Sesi", "Status Sesi", "Nilai",
        "Status Nilai", "Mulai", "Selesai", "Durasi (detik)",
        "Pelanggaran", "Waktu Jeda (detik)",
    ]

    def append_rows(sheet: Any, rows: Sequence[Mapping[str, Any]]) -> None:
        sheet.append(headers)
        for row in rows:
            score = row.get("score")
            status = "Belum dinilai"
            if score is not None:
                status = "Lulus" if float(score) >= passing_score else "Tidak lulus"
            sheet.append([
                row.get("user_id"),
                row.get("student_name") or "",
                row.get("student_class") or "",
                row.get("session_id"),
                row.get("status") or "",
                float(score) if score is not None else None,
                status,
                format_datetime(row.get("start_time")),
                format_datetime(row.get("end_time")),
                duration_seconds(row),
                int(row.get("violation_count") or 0),
                int(row.get("total_paused_seconds") or 0),
            ])

    append_rows(workbook.create_sheet("Hasil_Terbaru"), latest)
    append_rows(workbook.create_sheet("Semua_Percobaan"), all_attempts)

    metadata = workbook.create_sheet("Metadata_Ekspor")
    metadata.append(["Field", "Nilai"])
    metadata.append(["Waktu snapshot UTC", captured_at.isoformat()])
    metadata.append(["Status terminal", ", ".join(TERMINAL_STATUSES)])
    metadata.append(["Aturan hasil resmi", "Percobaan bernilai terbaru; fallback percobaan terminal terbaru"])
    metadata.append(["Jawaban mentah", "Tidak diekspor"])

    _style_workbook(workbook)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_question_pdf(
    exam: Mapping[str, Any],
    questions: Sequence[Mapping[str, Any]],
    options: Sequence[Mapping[str, Any]],
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.7 * cm, rightMargin=1.7 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ArchiveTitle", parent=styles["Title"], alignment=TA_CENTER, fontSize=15)
    body_style = ParagraphStyle("ArchiveBody", parent=styles["BodyText"], leading=15, spaceAfter=5)
    story: list[Any] = [
        Paragraph("NASKAH SOAL UJIAN", title_style),
        Spacer(1, 10),
    ]
    info_rows = [
        ["Judul", _clean_pdf_text(exam.get("exam_title"))],
        ["Mata Pelajaran", _clean_pdf_text(exam.get("subject") or "-")],
        ["Guru", _clean_pdf_text(exam.get("teacher_name"))],
        ["Jadwal", _clean_pdf_text(f"{format_datetime(exam.get('start_time'))} - {format_datetime(exam.get('end_time'))}")],
        ["Jumlah Soal", str(len(questions))],
    ]
    info_table = Table(info_rows, colWidths=[4 * cm, 12.5 * cm])
    info_table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.extend([info_table, Spacer(1, 18)])

    options_by_question: dict[int, list[Mapping[str, Any]]] = defaultdict(list)
    for option in options:
        options_by_question[int(option["question_id"])].append(option)

    for number, question in enumerate(questions, 1):
        story.append(Paragraph(f"<b>{number}. [{_clean_pdf_text(question.get('question_type'))}] ({float(question.get('points') or 0):g} poin)</b>", body_style))
        if question.get("stimulus"):
            story.append(Paragraph(f"<i>Stimulus:</i> {_clean_pdf_text(question['stimulus'])}", body_style))
        story.append(Paragraph(_clean_pdf_text(question.get("question_text")), body_style))
        for index, option in enumerate(options_by_question[int(question["question_id"])]):
            label = chr(65 + index) if index < 26 else str(index + 1)
            story.append(Paragraph(f"&nbsp;&nbsp;{label}. {_clean_pdf_text(option.get('option_text'))}", body_style))
        for key, label in (("image_url", "Gambar"), ("video_url", "Video"), ("audio_url", "Audio")):
            if question.get(key):
                story.append(Paragraph(f"<i>{label}:</i> {_clean_pdf_text(question[key])}", body_style))
        story.append(Spacer(1, 8))

    story.extend([PageBreak(), Paragraph("KUNCI JAWABAN", title_style), Spacer(1, 10)])
    for number, question in enumerate(questions, 1):
        key = answer_key(question, options_by_question[int(question["question_id"])]).replace("\n", "<br/>")
        story.append(Paragraph(f"<b>{number}.</b> {_clean_pdf_text(key).replace('&lt;br/&gt;', '<br/>')}", body_style))
    doc.build(story)
    return buffer.getvalue()


def create_results_pdf(
    exam: Mapping[str, Any],
    latest: Sequence[Mapping[str, Any]],
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    passing_score = float(exam.get("passing_score") or 70)
    summary = result_summary(latest, passing_score)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=1.2 * cm, rightMargin=1.2 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ResultTitle", parent=styles["Title"], alignment=TA_CENTER, fontSize=15)
    story: list[Any] = [
        Paragraph("LAPORAN HASIL UJIAN", title_style),
        Paragraph(_clean_pdf_text(exam.get("exam_title")), title_style),
        Spacer(1, 10),
    ]
    summary_rows = [
        ["Guru", _clean_pdf_text(exam.get("teacher_name")), "KKM", f"{passing_score:g}"],
        ["Peserta", summary["participants"], "Belum dinilai", summary["ungraded"]],
        ["Rata-rata", format_score(summary["average"]), "Lulus", summary["passed"]],
        ["Tertinggi", format_score(summary["highest"]), "Tidak lulus", summary["failed"]],
    ]
    summary_table = Table(summary_rows, colWidths=[3.3 * cm, 7 * cm, 3.3 * cm, 4 * cm])
    summary_table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.extend([summary_table, Spacer(1, 14)])

    rows: list[list[Any]] = [["No", "Nama", "Kelas", "Nilai", "Status", "Status Sesi", "Selesai", "Durasi", "Pelanggaran"]]
    if not latest:
        rows.append(["-", "Belum ada hasil ujian", "-", "-", "-", "-", "-", "-", "-"])
    for number, row in enumerate(latest, 1):
        score = row.get("score")
        grade_status = "Belum dinilai"
        if score is not None:
            grade_status = "Lulus" if float(score) >= passing_score else "Tidak lulus"
        duration = duration_seconds(row)
        rows.append([
            number,
            _clean_pdf_text(row.get("student_name")),
            _clean_pdf_text(row.get("student_class") or "-"),
            format_score(score),
            grade_status,
            _clean_pdf_text(row.get("status")),
            _clean_pdf_text(format_datetime(row.get("end_time"))),
            f"{duration or 0} dtk",
            int(row.get("violation_count") or 0),
        ])
    table = Table(rows, repeatRows=1, colWidths=[1 * cm, 5.3 * cm, 2.2 * cm, 2 * cm, 2.7 * cm, 2.5 * cm, 5 * cm, 2.2 * cm, 2.3 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
    ]))
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_artifact(path: Path, expected_sheets: Optional[Sequence[str]] = None) -> dict[str, Any]:
    if path.is_symlink() or not path.is_file():
        raise ArchiveError(f"Artifact tidak aman atau tidak ditemukan: {path}")
    if path.suffix == ".pdf" and not path.read_bytes().startswith(b"%PDF-"):
        raise ArchiveError(f"PDF tidak valid: {path}")
    if path.suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if expected_sheets and tuple(workbook.sheetnames) != tuple(expected_sheets):
                raise ArchiveError(f"Sheet XLSX tidak sesuai: {path}")
        finally:
            workbook.close()
    mode = stat.S_IMODE(path.stat().st_mode)
    if mode & 0o077:
        raise ArchiveError(f"Permission artifact terlalu terbuka: {path}")
    return {
        "path": path.as_posix(),
        "size_bytes": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def _relative_artifact(path: Path, root: Path, expected_sheets: Optional[Sequence[str]] = None) -> dict[str, Any]:
    metadata = validate_artifact(path, expected_sheets)
    metadata["path"] = path.relative_to(root).as_posix()
    return metadata


class TeacherExamArchiveExporter:
    def __init__(self, connection: Any, output: Path, captured_at: datetime):
        self.connection = connection
        self.output = output
        self.captured_at = captured_at

    async def _rows(self, sql: str, **params: Any) -> list[dict[str, Any]]:
        from sqlalchemy import text

        result = await self.connection.execute(text(sql), params)
        return [dict(row) for row in result.mappings().all()]

    async def selected_exams(self) -> list[dict[str, Any]]:
        return await self._rows(EXAM_SELECTION_SQL, captured_at=self.captured_at)

    async def count_data(self, exams: Sequence[Mapping[str, Any]]) -> dict[str, int]:
        from sqlalchemy import bindparam, text

        if not exams:
            return {"teachers": 0, "exams": 0, "questions": 0, "terminal_sessions": 0}
        exam_ids = [int(exam["exam_id"]) for exam in exams]
        count_stmt = text("""
            SELECT COUNT(*)
            FROM exam_sessions
            WHERE exam_id IN :exam_ids
              AND status IN ('submitted', 'completed')
        """).bindparams(bindparam("exam_ids", expanding=True))
        terminal_sessions = int((await self.connection.execute(count_stmt, {"exam_ids": exam_ids})).scalar_one())
        return {
            "teachers": len({int(exam["teacher_id"]) for exam in exams}),
            "exams": len(exams),
            "questions": sum(int(exam["question_count"]) for exam in exams),
            "terminal_sessions": terminal_sessions,
        }

    async def export(self, exams: Sequence[Mapping[str, Any]], database_identity: str) -> dict[str, Any]:
        artifacts: list[dict[str, Any]] = []
        warnings: list[dict[str, Any]] = []
        totals = {"teachers": 0, "exams": len(exams), "questions": 0, "terminal_sessions": 0, "latest_results": 0}
        grouped: dict[int, list[Mapping[str, Any]]] = defaultdict(list)
        for exam in exams:
            grouped[int(exam["teacher_id"])].append(exam)
        totals["teachers"] = len(grouped)

        for teacher_id, teacher_exams in grouped.items():
            first = teacher_exams[0]
            teacher_dir = self.output / f"guru_{teacher_id:06d}_{slugify(first['teacher_name'])}"
            _secure_mkdir(teacher_dir)
            teacher_path = teacher_dir / "info_guru.json"
            _write_json(teacher_path, {
                "id": teacher_id,
                "username": first.get("teacher_username"),
                "nama": first.get("teacher_name"),
                "role": first.get("teacher_role"),
                "aktif": first.get("teacher_is_active"),
                "jumlah_ujian": len(teacher_exams),
            })
            artifacts.append(_relative_artifact(teacher_path, self.output))

            for exam in teacher_exams:
                exam_id = int(exam["exam_id"])
                exam_dir = teacher_dir / f"ujian_{exam_id:06d}_{slugify(exam['exam_title'])}"
                _secure_mkdir(exam_dir)
                question_dir = exam_dir / "soal_dan_kunci"
                result_dir = exam_dir / "hasil_ujian"
                _secure_mkdir(question_dir)
                _secure_mkdir(result_dir)

                questions = await self._rows(QUESTION_SQL, exam_id=exam_id)
                options = await self._rows(OPTION_SQL, exam_id=exam_id)
                sessions = await self._rows(SESSION_SQL, exam_id=exam_id)
                latest = pick_latest_scored_sessions(sessions)
                totals["questions"] += len(questions)
                totals["terminal_sessions"] += len(sessions)
                totals["latest_results"] += len(latest)

                if not sessions:
                    warnings.append({"exam_id": exam_id, "kode": "tanpa_hasil", "pesan": "Belum ada sesi submitted/completed."})

                exam_info = dict(exam)
                exam_info["actual_question_count"] = len(questions)
                exam_info["terminal_session_count"] = len(sessions)
                exam_info["latest_result_count"] = len(latest)
                info_path = exam_dir / "info_ujian.json"
                _write_json(info_path, exam_info)
                artifacts.append(_relative_artifact(info_path, self.output))

                question_pdf = question_dir / "soal_dan_kunci.pdf"
                question_xlsx = question_dir / "soal_dan_kunci.xlsx"
                result_pdf = result_dir / "hasil_ujian.pdf"
                result_xlsx = result_dir / "hasil_ujian.xlsx"
                _write_bytes(question_pdf, create_question_pdf(exam, questions, options))
                _write_bytes(question_xlsx, create_question_xlsx(exam, questions, options))
                _write_bytes(result_pdf, create_results_pdf(exam, latest))
                _write_bytes(result_xlsx, create_results_xlsx(exam, latest, sessions, self.captured_at))
                artifacts.extend([
                    _relative_artifact(question_pdf, self.output),
                    _relative_artifact(question_xlsx, self.output, REQUIRED_QUESTION_SHEETS),
                    _relative_artifact(result_pdf, self.output),
                    _relative_artifact(result_xlsx, self.output, REQUIRED_RESULT_SHEETS),
                ])

        manifest = {
            "schema_version": 1,
            "captured_at_utc": self.captured_at,
            "database": database_identity,
            "criteria": {
                "owner_scope": {
                    "roles": ["teacher", "guruplus"],
                    "additional_usernames": ["kamad"],
                },
                "published": True,
                "finished_at_or_before_snapshot": True,
                "must_have_questions": True,
                "include_soft_deleted_exams": True,
                "terminal_statuses": list(TERMINAL_STATUSES),
                "include_answer_keys": True,
                "raw_student_answers_exported": False,
            },
            "totals": totals,
            "warnings": warnings,
            "dependencies": {
                name: _package_version(name) for name in ("SQLAlchemy", "asyncpg", "reportlab", "openpyxl")
            },
            "artifacts": artifacts,
        }
        manifest_path = self.output / "manifest.json"
        _write_json(manifest_path, manifest)
        return manifest


def _package_version(name: str) -> str:
    with suppress(importlib.metadata.PackageNotFoundError):
        return importlib.metadata.version(name)
    return "tidak tersedia"


def _preflight_output(target: Path, dry_run: bool) -> tuple[Path, Optional[Path]]:
    target = target.expanduser().resolve(strict=False)
    if target.is_symlink():
        raise ArchiveError(f"Target tidak boleh berupa symlink: {target}")
    if target.exists():
        if not target.is_dir() or any(target.iterdir()):
            raise ArchiveError(f"Target sudah ada dan tidak kosong: {target}")
    parent = target.parent
    if not parent.exists() or not parent.is_dir():
        if dry_run:
            return target, None
        parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    if parent.is_symlink():
        raise ArchiveError(f"Parent target tidak boleh berupa symlink: {parent}")
    if dry_run:
        return target, None
    usage = shutil.disk_usage(parent)
    if usage.free < 100 * 1024 * 1024:
        raise ArchiveError("Ruang kosong kurang dari 100 MiB.")
    staging = parent / f".{target.name}.partial-{os.getpid()}"
    if staging.exists() or staging.is_symlink():
        raise ArchiveError(f"Staging sudah ada: {staging}")
    _secure_mkdir(staging)
    return target, staging


async def run(args: argparse.Namespace) -> int:
    database_url = get_database_url()
    check_production_safety(database_url, args.allow_production_readonly)
    async_url = normalize_async_database_url(database_url)
    target, staging = _preflight_output(args.output, args.dry_run)
    if not args.dry_run:
        _ensure_export_dependencies()

    try:
        from sqlalchemy import text
        from sqlalchemy.ext.asyncio import create_async_engine
    except ImportError as exc:
        raise ArchiveError("SQLAlchemy async diperlukan untuk ekspor.") from exc

    engine = create_async_engine(async_url, pool_pre_ping=True)
    transaction = None
    try:
        async with engine.connect() as connection:
            transaction = await connection.begin()
            try:
                await connection.execute(text("SET TRANSACTION ISOLATION LEVEL REPEATABLE READ READ ONLY"))
                captured_at = (await connection.execute(text("SELECT CURRENT_TIMESTAMP"))).scalar_one()
                captured_at = _aware_datetime(captured_at) or datetime.now(timezone.utc)
                exporter = TeacherExamArchiveExporter(connection, staging or target, captured_at)
                exams = await exporter.selected_exams()
                counts = await exporter.count_data(exams)
                print(
                    "Data terpilih: "
                    f"{counts['teachers']} guru, {counts['exams']} ujian, "
                    f"{counts['questions']} soal, {counts['terminal_sessions']} sesi selesai."
                )
                if args.dry_run:
                    return 0
                await exporter.export(exams, redacted_database_identity(database_url))
            finally:
                if transaction and transaction.is_active:
                    await transaction.rollback()
    except Exception:
        if staging:
            print(f"Ekspor gagal. Staging dipertahankan untuk pemeriksaan: {staging}", file=sys.stderr)
        raise
    finally:
        await engine.dispose()

    if staging is None:
        raise ArchiveError("Staging tidak tersedia untuk ekspor.")
    staging.replace(target)
    print(f"Arsip selesai: {target}")
    return 0


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        return asyncio.run(run(args))
    except (ArchiveError, OSError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    except KeyboardInterrupt:
        print("Dibatalkan.", file=sys.stderr)
        return 130


if __name__ == "__main__":
    raise SystemExit(main())
