from __future__ import annotations

import importlib.util
import json
import stat
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "export_teacher_exam_archive.py"
SPEC = importlib.util.spec_from_file_location("export_teacher_exam_archive", SCRIPT_PATH)
assert SPEC and SPEC.loader
archive = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(archive)


def _exam() -> dict:
    return {
        "exam_id": 12,
        "exam_title": "Ujian Akhir / IPA",
        "subject": "IPA",
        "exam_type": "UAS",
        "academic_year": "2025/2026",
        "teacher_name": "Guru Contoh",
        "start_time": datetime(2026, 8, 1, 1, tzinfo=timezone.utc),
        "end_time": datetime(2026, 8, 1, 3, tzinfo=timezone.utc),
        "duration_minutes": 120,
        "question_count": 1,
        "passing_score": 70,
        "is_deleted": True,
    }


def _question() -> dict:
    return {
        "question_id": 21,
        "order_index": 1,
        "question_type": "multiple_choice",
        "question_subtype": None,
        "pgk_type": None,
        "difficulty_level": "medium",
        "points": 10,
        "stimulus": "Baca <b>baik-baik</b>",
        "question_text": "Hasil 1 < 2 adalah?",
        "question_settings": {"acceptable_answers": ["benar"], "internal_secret": "jangan"},
        "image_url": "https://example.invalid/image.png",
        "video_url": None,
        "audio_url": None,
    }


def _options() -> list[dict]:
    return [
        {
            "option_id": 31,
            "question_id": 21,
            "option_text": "Benar",
            "is_correct": True,
            "order_index": 1,
            "option_group": "standard",
            "pair_id": None,
            "option_metadata": {},
        },
        {
            "option_id": 32,
            "question_id": 21,
            "option_text": "Salah",
            "is_correct": False,
            "order_index": 2,
            "option_group": "standard",
            "pair_id": None,
            "option_metadata": {},
        },
    ]


def _session(session_id: int, user_id: int, score, ended_minutes: int) -> dict:
    end = datetime(2026, 8, 2, 4, tzinfo=timezone.utc) + timedelta(minutes=ended_minutes)
    return {
        "session_id": session_id,
        "user_id": user_id,
        "student_name": f"Siswa {user_id}",
        "student_class": "X-A",
        "start_time": end - timedelta(minutes=45),
        "end_time": end,
        "status": "submitted",
        "score": score,
        "violation_count": 1,
        "total_paused_seconds": 0,
    }


def test_exam_selection_includes_teacher_guruplus_and_kamad_published_finished():
    normalized = " ".join(archive.EXAM_SELECTION_SQL.lower().split())
    assert "lower(trim(u.role)) in ('teacher', 'guruplus')" in normalized
    assert "lower(trim(u.username)) = 'kamad'" in normalized
    assert "e.is_published is true" in normalized
    assert "e.end_time <= :captured_at" in normalized
    assert "join questions q" in normalized
    assert "is_deleted" not in normalized.split("where", 1)[1].split("group by", 1)[0]
    assert "insert " not in normalized
    assert "update " not in normalized
    assert "delete " not in normalized


def test_terminal_queries_only_use_submitted_and_completed():
    assert "status in ('submitted', 'completed')" in " ".join(archive.SESSION_SQL.lower().split())
    assert archive.TERMINAL_STATUSES == ("submitted", "completed")


def test_pick_latest_scored_prefers_latest_scored_and_falls_back_latest_terminal():
    rows = [
        _session(5, 1, None, 50),
        _session(4, 1, 88, 40),
        _session(3, 1, 70, 30),
        _session(8, 2, None, 80),
        _session(7, 2, None, 70),
    ]
    selected = archive.pick_latest_scored_sessions(rows)
    selected_by_user = {row["user_id"]: row for row in selected}
    assert selected_by_user[1]["session_id"] == 4
    assert selected_by_user[2]["session_id"] == 8


def test_slugify_blocks_traversal_and_keeps_id_names_stable():
    assert archive.slugify("../../Guru / A") == "guru_a"
    assert "/" not in archive.slugify("A/B")
    assert ".." not in archive.slugify("../..")
    assert archive.slugify("") == "tanpa_nama"


def test_answer_key_only_keeps_supported_settings():
    rendered = archive.answer_key(_question(), _options())
    assert "Benar" in rendered
    assert "acceptable_answers" in rendered
    assert "internal_secret" not in rendered


def test_question_xlsx_is_real_workbook_with_expected_sheets(tmp_path):
    path = tmp_path / "questions.xlsx"
    archive._write_bytes(path, archive.create_question_xlsx(_exam(), [_question()], _options()))
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        assert tuple(workbook.sheetnames) == archive.REQUIRED_QUESTION_SHEETS
        assert workbook["Soal"].max_row == 2
        assert workbook["Pilihan"].max_row == 3
        key_text = workbook["Kunci_Jawaban"]["D2"].value
        assert "Benar" in key_text
        assert "internal_secret" not in key_text
    finally:
        workbook.close()


def test_results_xlsx_has_latest_and_all_attempts_without_sensitive_fields(tmp_path):
    rows = [_session(5, 1, None, 50), _session(4, 1, 88, 40)]
    latest = archive.pick_latest_scored_sessions(rows)
    path = tmp_path / "results.xlsx"
    archive._write_bytes(
        path,
        archive.create_results_xlsx(_exam(), latest, rows, datetime.now(timezone.utc)),
    )
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        assert tuple(workbook.sheetnames) == archive.REQUIRED_RESULT_SHEETS
        assert workbook["Hasil_Terbaru"].max_row == 2
        assert workbook["Semua_Percobaan"].max_row == 3
        headers = [cell.value for cell in workbook["Semua_Percobaan"][1]]
        rendered_headers = " ".join(str(value).lower() for value in headers)
        for forbidden in ("answer_text", "selected", "ip address", "user agent", "password", "token"):
            assert forbidden not in rendered_headers
    finally:
        workbook.close()


def test_empty_results_are_valid_in_workbook():
    payload = archive.create_results_xlsx(_exam(), [], [], datetime.now(timezone.utc))
    assert payload.startswith(b"PK")


def test_result_summary_does_not_turn_ungraded_into_zero():
    rows = [_session(1, 1, None, 0), _session(2, 2, 80, 1)]
    summary = archive.result_summary(rows, passing_score=70)
    assert summary["participants"] == 2
    assert summary["scored"] == 1
    assert summary["ungraded"] == 1
    assert summary["average"] == 80
    assert summary["lowest"] == 80


def test_output_preflight_refuses_nonempty_target(tmp_path):
    target = tmp_path / "MT2026-08-10"
    target.mkdir()
    (target / "existing.txt").write_text("jangan timpa", encoding="utf-8")
    with pytest.raises(archive.ArchiveError, match="tidak kosong"):
        archive._preflight_output(target, dry_run=False)


def test_output_preflight_allows_existing_empty_target(tmp_path):
    target = tmp_path / "MT2026-08-10"
    target.mkdir()
    resolved, staging = archive._preflight_output(target, dry_run=False)
    assert resolved == target.resolve()
    assert staging is not None and staging.is_dir()


def test_dry_run_preflight_does_not_create_parent(tmp_path):
    target = tmp_path / "missing" / "MT2026-08-10"
    resolved, staging = archive._preflight_output(target, dry_run=True)
    assert resolved == target.resolve(strict=False)
    assert staging is None
    assert not target.parent.exists()


def test_validate_artifact_hash_and_permissions(tmp_path):
    path = tmp_path / "info.json"
    archive._write_json(path, {"ok": True})
    metadata = archive.validate_artifact(path)
    assert len(metadata["sha256"]) == 64
    assert stat.S_IMODE(path.stat().st_mode) == 0o600
    assert json.loads(path.read_text(encoding="utf-8")) == {"ok": True}


def test_database_identity_redacts_credentials():
    identity = archive.redacted_database_identity(
        "postgresql+asyncpg://secret_user:secret_password@db.example:5432/siab1"
    )
    assert identity == "postgresql://db.example:5432/siab1"
    assert "secret" not in identity


def test_pdf_text_escapes_markup():
    cleaned = archive._clean_pdf_text("1 < 2 & <b>aman</b>")
    assert "&lt;" in cleaned
    assert "&amp;" in cleaned
    assert "<b>" not in cleaned


def test_pdf_renderers_when_reportlab_available():
    pytest.importorskip("reportlab")
    question_pdf = archive.create_question_pdf(_exam(), [_question()], _options())
    result_pdf = archive.create_results_pdf(_exam(), [])
    assert question_pdf.startswith(b"%PDF-")
    assert result_pdf.startswith(b"%PDF-")
